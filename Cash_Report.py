
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
from fpdf import FPDF
from num2words import num2words
from tools import open_con
import os
from Validation import validate_CNIC, validate_date, convert_date_format, validate_string


class Cash_Report:
    def __init__(self):
        self.__root = Tk()
        self.__font = ('Cambria', 14, 'bold')
        self.__root.geometry('1400x700+50+50')
        self.__root.title("Cash Report")
        self.btn_font = ('Cambria', 12)
        self.filter = {}
        self.rec_id = 0
        self.total_fee = 0
        self.agr_data = None
        # self.__root.tk.call('lappend', 'auto_path',
        #                     r'C:\Users\Hamid Shah\Desktop\newtheam\awthemes-10.4.0')
        # self.__root.tk.call('package', 'require', 'awdark')
        my_style = ttk.Style(self.__root)
        # my_style.theme_use('awdark')

        self.top_frame = Frame(self.__root, relief=RIDGE, borderwidth=2)
        self.top_frame.pack(side=TOP, fill=X)
        self.top_frame1 = Frame(self.__root, relief=RIDGE, borderwidth=2)
        self.top_frame1.pack(side=TOP, fill=X)

        self.filter_type = Listbox(self.top_frame1, height=1,
                                   font=self.btn_font, exportselection=0)
        self.filter_type.grid(row=1, column=1, padx=10, pady=10)
        self.filter_type.insert(0, 'Domicile Date')
        self.filter_type.insert(1, 'Applicant Name')
        self.filter_type.insert(2, 'CNIC')
        self.filter_type.insert(3, 'Payment Type')
        self.filter_type.insert(4, 'Govt_Fee')
        self.filter_type.insert(5, 'Application Type')
        self.filter_type.insert(6, 'Duplicate Entry')
        self.filter_type.select_set(0)

        self.filter_input = Entry(self.top_frame1, font=self.btn_font)
        self.filter_input.grid(row=1, column=2, padx=10, pady=10)
        self.left_frame = Frame(self.__root, relief=RIDGE, borderwidth=2)
        self.left_frame.pack(side=LEFT, fill=BOTH)
        self.right_frame = Frame(self.__root, relief=RIDGE, borderwidth=2)
        self.right_frame.pack(side=LEFT, fill=BOTH)
        self.bottom_frame = Frame(
            self.right_frame, relief=RIDGE, borderwidth=2)
        self.bottom_frame.pack(side=BOTTOM, fill=X)
        self.status_label = Entry(
            self.bottom_frame, text='Status Bar', width=280, relief='flat', font=self.btn_font)
        self.status_label.grid(row=1, column=0, sticky=W)
        self.title_label = Label(
            self.top_frame, text='Domicile Cash Report', font=self.__font)
        self.title_label.pack()

        self.excel_btn = Button(self.left_frame, text='Read Excel', command=self.read_excel,
                                width=15, font=self.btn_font)
        self.excel_btn.grid(row=1, column=0, padx=10, pady=10, sticky=W)
        self.cash_report_btn = Button(self.left_frame, text='Cash Report', command=self.cash_report_pdf,
                                      width=15,  font=self.btn_font)
        self.cash_report_btn.grid(row=2, column=0, padx=10, pady=10, sticky=W)
        self.notesheet_btn = Button(self.left_frame, text='Note Sheet', command=self.domicile_note_sheet,
                                    width=15,  font=self.btn_font)
        self.notesheet_btn.grid(row=3, column=0, padx=10, pady=10, sticky=W)
        self.challan_list_btn = Button(self.left_frame, text='Challan List', command=self.domicile_challan_list,
                                       width=15,  font=self.btn_font)
        self.challan_list_btn.grid(row=4, column=0, padx=10, pady=10, sticky=W)
        self.challan_btn = Button(self.left_frame, text='Challan', command=self.domicile_challan,
                                  width=15,  font=self.btn_font)
        self.challan_btn.grid(row=5, column=0, padx=10, pady=10, sticky=W)

        my_style.configure('Treeview', background="lightblue",
                           rowheight=36, font=self.btn_font)
        my_style.configure("Treeview.Heading",
                           font=self.__font, background="blue")
        self.trv = ttk.Treeview(
            self.right_frame, selectmode='browse', height=7)
        self.trv.pack(fill=BOTH, expand=TRUE)
        # , "7", "8", "9")
        self.trv["columns"] = ("1", "2", "3", "4", "5", "6", "7", "8", "9")
        self.trv['show'] = 'headings'
        self.trv.column("1", width=50, anchor='center')
        self.trv.column("2", width=150, anchor='center')
        self.trv.column("3", width=130, anchor='w')
        self.trv.column("4", width=250, anchor='w')
        self.trv.column("5", width=130, anchor='w')
        self.trv.column("6", width=130, anchor='center')
        self.trv.column("7", width=150, anchor='center')
        self.trv.column("8", width=150, anchor='center')
        self.trv.column("9", width=150, anchor='center')

        self.trv.heading("1", text="ID", anchor='center')
        self.trv.heading("2", text="Domicile Date", anchor='center')
        self.trv.heading("3", text="CNIC", anchor='w')
        self.trv.heading("4", text="Applicant Name", anchor='w')
        self.trv.heading("5", text="Payment Type", anchor='w')
        self.trv.heading("6", text="Govt Fee", anchor='center')
        self.trv.heading("7", text="Application Type", anchor='w')
        self.trv.heading("8", text="Request Type", anchor='center')
        self.trv.heading("9", text="Duplicate Entry", anchor='center')

        self.trv.bind("<Button 1>", self.trv_click)
        self.add_filter_btn = Button(self.top_frame1, text='Add Filter', command=self.add_filter,
                                     font=self.btn_font, width=15)
        self.add_filter_btn.grid(row=1, column=3, padx=10, pady=10)
        self.apply_filter_btn = Button(self.top_frame1, text='Apply Filter', command=self.apply_filter,
                                       font=self.btn_font, width=15)
        self.apply_filter_btn.grid(row=1, column=4, padx=10, pady=10)
        self.clear_filter_btn = Button(self.top_frame1, text='Clear Filter', command=self.clear_filter,
                                       font=self.btn_font, width=15)
        self.clear_filter_btn.grid(row=1, column=5, padx=10, pady=10)
        self.defualt_filter_btn = Button(self.top_frame1, text='Defualt Filter', command=self.defualt_filter,
                                         font=self.btn_font, width=15)
        self.defualt_filter_btn.grid(row=1, column=6, padx=10, pady=10)
        self.new_btn = Button(self.top_frame1, text='New Record', command= lambda:self.edit_record('new'),
                               font=self.btn_font, width=15)
        self.new_btn.grid(row=1, column=7, padx=10, pady=10)
        self.edit_btn = Button(self.top_frame1, text='Edit Record', command= lambda:self.edit_record('edit'),
                               font=self.btn_font, width=15)
        self.edit_btn.grid(row=1, column=8, padx=10, pady=10)
        self.exit_btn = Button(self.top_frame1, text='Exit', command=self.__root.destroy,
                               font=self.btn_font, width=15)
        self.exit_btn.grid(row=1, column=9, padx=10, pady=10)
        self.get_cur_server_date()

        self.filter_input.insert(0, self.current_date)
        self.defualt_filter()

    def check_date(self, report_date, cur):
        date_dict = convert_date_format('dd/mm/yyyy', '/', report_date)
        my_date_format = date_dict['yyyy'] + '-' + \
            date_dict['mm'] + '-' + date_dict['dd']
        Query = "Select Domicile_Date from Cash_Report Where Domicile_Date = %s;"
        cur.execute(Query, [my_date_format])
        data = cur.fetchall()
        if data is None:
            return 'Date not exist'
        elif len(data) == 0:
            return 'Date not exist'
        else:
            return 'Date already exist'

    def read_excel(self):
        filename = filedialog.askopenfilename()
        if len(filename) == 0:
            return messagebox.showerror("Error", "You did not select any file")
        wb = load_workbook(filename)
        # wb = load_workbook(f'C:\\Users\\Hamid Shah\\Downloads\\domicile.xlsx')
        ws = wb['Worksheet']
        a = 0
        rec_counter = 0
        cnic_list = []
        # f = open('test.txt', 'w')
        con, cur = open_con(False)
        for row in ws.values:
            a = a + 1
            if a != 1:  # scape first row of excel file which is header row
                if a == 2:  # first data row to check date in db
                    if self.check_date(row[0], cur) == 'Date already exist':
                        wb.close()
                        cur.close()
                        con.close()
                        return messagebox.showerror('Error', 'Data for this file already uploaded')
                if row[2] not in cnic_list:
                    dup_value = 'No'
                    cnic_list.append(row[2].strip())
                else:
                    dup_value = 'Yes'

                if row[5] is None:
                    Govt_Fee = "0"
                    Payment_Type = "Not paid"
                    app_type = 'online'
                elif row[5].strip() == 'Cash':
                    Govt_Fee = "200"
                    Payment_Type = "Cash"
                    app_type = 'offline'
                else:
                    Govt_Fee = "Paid in Bank"
                    Payment_Type = "Challan"
                    app_type = 'offline'
                if row[6] == 'Revised' or row[6] == 'Duplicate':
                    Govt_Fee = "0"
                    Payment_Type = "Not paid"
                date_dict = convert_date_format('dd/mm/yyyy', '/', row[0])
                my_date_format = date_dict['yyyy'] + '-' + \
                    date_dict['mm'] + '-' + date_dict['dd']
                Query = "Insert into Cash_Report (Domicile_Date, Applicant_Name, cnic, Application_Type, Duplicate_Entry, Govt_Fee, Payment_Type, Domicile_No, Request_Type) values (%s, %s, %s, %s, %s, %s, %s, %s, %s);"
                parm_list = [my_date_format, row[1], row[2].strip().strip("'"), app_type, dup_value, Govt_Fee, Payment_Type, row[7][row[7].find('-', 6)+1:], row[6]]
                rec_counter += 1
                cur.execute(Query, parm_list)
                print(Query)
                con.commit()
                self.status_label.delete(0, 'end')
                self.status_label.insert(
                    0, 'Total Records Inserted:- {}'.format(rec_counter))

        # f.close()
        cur.close()
        con.close()
        wb.close()

    def get_cur_server_date(self):
        con, cur = open_con(False)
        
        cur.execute("SELECT curdate();")
        data = cur.fetchone()
        self.current_date = data[0]
        cur.close()
        con.close()

    def add_filter(self):
        # making sure both filter type and filter inpute must provide some input
        if len(self.filter_type.curselection()) != 0 and len(self.filter_input.get()) != 0:
            if self.filter_type.get(self.filter_type.curselection()) == 'Domicile Date':
                # validation for date
                valid_result = validate_date(self.filter_input.get())
                if valid_result != 'valid':
                    return messagebox.showerror('Error', valid_result)
                else:
                    self.filter['Domicile_Date'] = self.filter_input.get()
            elif self.filter_type.get(self.filter_type.curselection()) == 'CNIC':
                valid_result = validate_CNIC(self.filter_input.get())
                if valid_result != 'valid':
                    return messagebox.showerror('Error', valid_result)
                else:
                    self.filter['CNIC'] = self.filter_input.get()
            elif self.filter_type.get(self.filter_type.curselection()) == 'Applicant Name':
                valid_result = validate_string(self.filter_input.get())
                if valid_result != 'valid':
                    return messagebox.showerror('Error', valid_result)
                else:
                    self.filter['Applicant_Name'] = self.filter_input.get()
            elif self.filter_type.get(self.filter_type.curselection()) == 'Payment Type':
                if self.filter_input.get() == 'Challan' or self.filter_input.get() == 'Cash':
                    self.filter['Payment_Type'] = self.filter_input.get()
                else:
                    return messagebox.showerror('Error', 'Input should be Challan or Cash')

            elif self.filter_type.get(self.filter_type.curselection()) == 'Application Type':
                if self.filter_input.get() == 'offline' or self.filter_input.get() == 'online':
                    self.filter['Application_Type'] = self.filter_input.get()
                else:
                    return messagebox.showerror('Error', 'Input should be offline or online')

            elif self.filter_type.get(self.filter_type.curselection()) == 'Duplicate Entry':
                if self.filter_input.get() == 'Yes' or self.filter_input.get() == 'No':
                    self.filter['Duplicate_Entry'] = self.filter_input.get()
                else:
                    return messagebox.showerror('Error', 'Input should be "Yes" or "No"')
            elif self.filter_type.get(self.filter_type.curselection()) == 'Govt_Fee':
                if self.filter_input.get() == 'Paid in Bank' or self.filter_input.get() == '200':
                    self.filter['Govt_Fee'] = self.filter_input.get()
                else:
                    return messagebox.showerror('Error', 'Input should be "200" or "Paid in Bank"')
            elif self.filter_type.get(self.filter_type.curselection()) == 'Request Type':
                if self.filter_input.get() == 'New' or self.filter_input.get() == 'Revised' or self.filter_input.get() == 'Duplicate':
                    self.filter['Request_Type'] = self.filter_input.get()
                else:
                    return messagebox.showerror('Error', 'Input should be "New", "Duplicate" or "Revised"')
            self.status_label.delete(0, 'end')
            self.status_label.insert(0, 'Filter = {}'.format(self.filter))

    def defualt_filter(self):
        self.filter = {}
        self.filter['Domicile_Date'] = "{}".format(self.current_date)
        self.filter['Duplicate_Entry'] = "No"
        self.filter['Application_Type'] = "offline"
        self.filter['Request_Type'] = "New"
        self.status_label.delete(0, 'end')
        self.status_label.insert(0, 'Filter = {}'.format(self.filter))

    def clear_filter(self):
        self.filter = {}
        self.status_label.delete(0, 'end')
        self.status_label.insert(0, 'Filter Cleared')

    def apply_filter(self):
        Query = "Select Domicile_ID, Domicile_Date, CNIC, Applicant_Name, Payment_Type, Application_Type, Request_Type, Govt_Fee, Duplicate_Entry from Cash_Report"
        lp = 0
        for item in self.filter:
            lp = lp + 1
            if lp == 1:
                Query = Query + \
                    " Where {} = '{}'".format(item, self.filter.get(item))
            else:
                Query = Query + \
                    " And {} = '{}'".format(item, self.filter.get(item))
        Query = Query + ";"

        self.trv.delete(*self.trv.get_children())
        con, cur = open_con(True)
        cur.execute(Query)
        self.agr_data = cur.fetchall()
        cur.close()
        con.close()
        self.status_label.delete(0, 'end')
        self.status_label.insert(
            0, 'Total Records:- {}'.format(len(self.agr_data)))
        for row in self.agr_data:
            self.trv.insert("", 'end', values=(
                row['Domicile_ID'], row['Domicile_Date'], row['CNIC'], row['Applicant_Name'], row['Payment_Type'], row['Govt_Fee'], row['Application_Type'], row['Request_Type'], row['Duplicate_Entry']))

    def trv_click(self, event):
        for item in self.trv.selection():
            selected_item = self.trv.item(item, 'values')
            self.rec_id = selected_item[0]

    def cash_report_pdf(self):

        if self.agr_data is None:
            return messagebox.showerror('Error', 'No Data in Grid')

        pdf = FPDF()
        pdf.add_page()

        pdf.set_font('courier', 'B', size=20)

        pdf.ln(7)
        #pdf.cell(20, 6, txt='', align='C')
        pdf.set_fill_color(152, 163, 155)
        pdf.cell(0, 10, txt='Domicile Cash Report of {}'.format(self.agr_data[0]['Domicile_Date']),
                 ln=1, align='L', fill=True)
        pdf.ln(6)
        pdf.set_font('courier', 'B', size=14)

        sl = 0
        self.total_fee = 0
        pdf.cell(17, 6, txt='S.No.', border=1)
        pdf.cell(90, 6, txt='Applicant Name', border=1)
        pdf.cell(40, 6, txt='Payment Type', border=1)
        pdf.cell(40, 6, txt='Govt Fee', ln=1, border=1)
        pdf.set_font('courier', size=14)
        for row in self.agr_data:
            sl += 1
            pdf.cell(17, 6, txt='{}'.format(sl), border=1)
            pdf.cell(90, 6, txt='{}'.format(
                row['Applicant_Name']), border=1)
            pdf.cell(40, 6, txt='{}'.format(
                row['Payment_Type']), border=1)
            pdf.cell(40, 6, txt='{}'.format(
                row['Govt_Fee']), ln=1, border=1)
            if row['Govt_Fee'].strip() == '200':
                self.total_fee = self.total_fee + 200
        print(self.total_fee)
        pdf.set_font('courier', 'B', size=14)
        pdf.cell(147, 6, txt='Total:-', align='R', border=1)
        pdf.cell(40, 6, txt='{}'.format(self.total_fee), ln=1, border=1)
        pdf.ln(12)
        pdf.ln(6)
        pdf.cell(170, 6, txt='Domicile Clerk', align='R', ln=1)
        pdf.set_font('courier', 'BU', size=14)
        pdf.cell(50, 6, txt='Accounts Assistant', align='L', ln=1)
        pdf.ln(6)

        pdf.output('Cash_Report.pdf')

        path = 'Cash_Report.pdf'
        os.system(path)

    def domicile_note_sheet(self):
        if self.agr_data is None:
            return messagebox.showerror('Error', 'No Data in Grid')

        pdf = FPDF()
        pdf.add_page()

        pdf.set_font('courier', 'B', size=20)
        pdf.set_fill_color(211, 211, 211)
        #pdf.image('govt_logo.png', x=10, y=10, w=30, h=30)
        pdf.ln(7)
        #pdf.cell(20, 6, txt='', align='C')
        pdf.set_fill_color(152, 163, 155)
        pdf.cell(0, 10, txt='Domicile Note Sheet of {}'.format(self.agr_data[0]['Domicile_Date']),
                 ln=1, align='L', fill=True)
        pdf.ln(6)
        pdf.set_font('courier', size=14)
        sl = 0
        total_fee = 0
        dup_counter = 0
        online_counter = 0
        new_counter = 0
        revised_counter = 0
        duplicate_dom = 0
        pdf.set_font('courier', 'B', size=14)
        pdf.cell(20, 6, txt='S.No.', border=1)
        pdf.cell(120, 6, txt='Applicant Name', border=1)
        pdf.cell(40, 6, txt='Request Type', border=1, ln=1)
        pdf.set_font('courier', size=14)
        for row in self.agr_data:
            sl += 1
            pdf.cell(20, 6, txt='{}'.format(sl), border=1)
            pdf.cell(120, 6, txt='{}'.format(
                row['Applicant_Name']), border=1)
            pdf.cell(40, 6, txt='{}'.format(
                row['Request_Type']), border=1, ln=1)
            if row['Duplicate_Entry'] == 'Yes':
                dup_counter += 1
            if row['Application_Type'] == 'online':
                online_counter += 1
            if row['Request_Type'] == 'New':
                new_counter += 1
            elif row['Request_Type'] == 'Revised':
                revised_counter += 1
            elif row['Request_Type'] == 'Duplicate':
                duplicate_dom += 1

            if row['Govt_Fee'] == '200':
                total_fee = total_fee + 200
        pdf.set_font('courier', 'BU', size=14)
        pdf.ln(6)
        pdf.cell(40, 6, txt='Summery', align='L', ln=1)
        pdf.set_font('courier', size=14)
        pdf.cell(40, 6, txt='Total Online Entries:- {}'.format(online_counter), ln=1)
        pdf.cell(40, 6, txt='Total Duplicate Entries:- {}'.format(dup_counter), ln=1)
        pdf.cell(40, 6, txt='New Domiciles:- {}'.format(new_counter), ln=1)
        pdf.cell(40, 6, txt='Duplicate Domiciles:- {}'.format(duplicate_dom), ln=1)
        pdf.cell(40, 6, txt='Revised Domiciles:- {}'.format(revised_counter), ln=1)
        pdf.cell(40, 6, txt='Total Entries:- {}'.format(len(self.agr_data)), ln=1)
        pdf.ln(12)
        pdf.ln(6)
        pdf.cell(170, 6, txt='Domicile Clerk', align='R', ln=1)
        pdf.set_font('courier', 'BU', size=14)
        pdf.cell(50, 6, txt='Accounts Assistant', align='L', ln=1)
        pdf.ln(6)

        pdf.output('Domicile_Note_Sheet.pdf')

        path = 'Domicile_Note_Sheet.pdf'
        os.system(path)

    def domicile_challan_list(self):
        if self.agr_data is None:
            return messagebox.showerror('Error', 'No Data in Grid')

        pdf = FPDF()
        pdf.add_page()

        pdf.set_font('courier', 'B', size=20)

        #pdf.image('govt_logo.png', x=10, y=10, w=30, h=30)
        pdf.ln(7)
        #pdf.cell(20, 6, txt='', align='C')

        pdf.set_fill_color(152, 163, 155)
        pdf.cell(0, 10, txt='Domicile Note Sheet of {}'.format(self.agr_data[0]['Domicile_Date']),
                 ln=1, align='L', fill=True)
        pdf.ln(6)
        pdf.set_font('courier', size=14)
        sl = 0
        self.total_fee = 0
        dup_counter = 0
        online_counter = 0
        for row in self.agr_data:
            sl += 1
            pdf.cell(15, 6, txt='{}'.format(sl), border=1)
            pdf.cell(90, 6, txt='{}'.format(
                row['Applicant_Name']), border=1, ln=1)
            if row['Duplicate_Entry'] == 'Yes':
                dup_counter += 1
            if row['Application_Type'] == 'online':
                online_counter += 1
            if row['Govt_Fee'] == '200':
                self.total_fee = self.total_fee + 200

        pdf.ln(6)

        pdf.output('Domicile_Note_Sheet.pdf')

        path = 'Domicile_Note_Sheet.pdf'
        os.system(path)

    def domicile_challan(self):
        self.total_fee = 0
        if self.agr_data is None:
            return messagebox.showerror('Error', 'No Data in Grid')
        for item in self.agr_data:
            if item['Govt_Fee'] == '200':
                self.total_fee += 200
        if self.total_fee == 0:
            return messagebox.showerror('Error', 'Fee for current filter is not due')
        challan_date = self.agr_data[0]['Domicile_Date']
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()

        pdf.set_font('arial', size=11)

        pdf.cell(60, 6, txt='PROVINCIAL', align='L')
        pdf.cell(70, 6, txt='Treasury or Sub-Treasury', align='C')
        pdf.cell(60, 6, txt='Form-32/A', align='C', ln=1)

        pdf.cell(60, 6, txt='Challan of Cash Paid in to', align='L')
        pdf.cell(70, 6, txt='National Bank of Pakistan', align='C')
        pdf.cell(60, 6, txt='', align='C', ln=1)

        pdf.cell(60, 6, txt='Challan No.', align='L')
        pdf.cell(70, 6, txt='State Bank of Pakistan', align='C')
        pdf.cell(60, 6, txt='', align='C', ln=1)
        pdf.ln(4)
        pdf.cell(35, 6, txt='By whom Tender', align='L', border='LTR')
        pdf.cell(65, 6, txt='To be filled in the Remittes', align='C', border=1)
        pdf.cell(35, 6, txt='Amount', align='C', border='LTRB')
        pdf.cell(60, 6, txt='To be filled by the  the Treasury',
                 align='L', border='LTR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='Names or', align='C', border='LR')
        pdf.cell(35, 6, txt='Full Particulars ', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(60, 6, txt='Departmental Office of',
                 align='L',  border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='designation', align='C', border='LR')
        pdf.cell(35, 6, txt='of the ', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(60, 6, txt='the Treasury', align='L', border='LBR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='and address', align='C', border='LR')
        pdf.cell(35, 6, txt='remittance and', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='Head of', align='C', border='LR')
        pdf.cell(30, 6, txt='Order of', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='of the person', align='C', border='LR')
        pdf.cell(35, 6, txt='of authority (if', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='Account', align='C', border='LR')
        pdf.cell(30, 6, txt='the Bank', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='of whom behalf', align='C', border='LR')
        pdf.cell(35, 6, txt='any) ', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LBR')
        pdf.cell(30, 6, txt='money is paid', align='C', border='LBR')
        pdf.cell(35, 6, txt='', align='C', border='LBR')
        pdf.cell(35, 6, txt='', align='C', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR', ln=1)

        pdf.cell(35, 6, txt='Domicile Branch', align='L', border='LR')
        pdf.cell(30, 6, txt='District', align='C', border='LR')
        pdf.cell(35, 6, txt='Domicile Fee', align='C', border='LR')
        pdf.cell(35, 6, txt='{}'.format(
            self.total_fee), align='C', border='LR')
        pdf.cell(30, 6, txt='C-03806', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='DC Office, ICT', align='L', border='LR')
        pdf.cell(30, 6, txt='Magistrate', align='C', border='LR')
        pdf.cell(35, 6, txt='as per', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='Islamabad', align='L', border='LR')
        pdf.cell(30, 6, txt='Islamabad', align='C', border='LR')
        pdf.cell(35, 6, txt='attached list', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='{}'.format(challan_date), align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)
        pdf.cell(35, 6, txt='', align='L', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR')
        pdf.cell(35, 6, txt='', align='C', border='LBR')
        pdf.cell(35, 6, txt='', align='C', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR', ln=1)
        pdf.set_font('arial', 'B', size=11)
        pdf.cell(35, 8, txt='', align='L', border='LBR')
        pdf.cell(30, 8, txt='', align='C', border='LBR')
        pdf.cell(35, 8, txt='Total', align='C', border='LBR')
        pdf.cell(35, 8, txt='{}'.format(
            self.total_fee), align='C', border='LBR')
        pdf.cell(30, 8, txt='', align='C', border='LBR')
        pdf.cell(30, 8, txt='', align='C', border='LBR', ln=1)

        pdf.set_font('arial', size=11)
        pdf.cell(65, 10, txt='In words Rupees ', align='R')
        pdf.set_font('arial', 'B', size=11)
        pdf.cell(0, 10, txt='{} Only'.format(
            num2words(self.total_fee)), align='L', ln=1)
        pdf.ln(5)
        pdf.set_font('arial', size=11)
        pdf.cell(50, 5, txt='Signature', align='LC')
        pdf.cell(50, 5, txt='Amount', align='C')
        pdf.cell(50, 5, txt='Received', align='C')
        pdf.cell(50, 5, txt='Treasury Officer', align='C', ln=1)

        pdf.cell(50, 5, txt='', align='LC')
        pdf.cell(50, 5, txt='', align='C')
        pdf.cell(50, 5, txt='Payment Treasury', align='C')
        pdf.cell(50, 5, txt='', align='C', ln=1)
        pdf.ln(2)
        pdf.cell(0, 1, txt='', border='T', ln=1)
        # second page
        pdf.ln(2)
        pdf.set_font('arial', size=11)

        pdf.cell(60, 6, txt='PROVINCIAL', align='L')
        pdf.cell(70, 6, txt='Treasury or Sub-Treasury', align='C')
        pdf.cell(60, 6, txt='Form-32/A', align='C', ln=1)

        pdf.cell(60, 6, txt='Challan of Cash Paid in to', align='L')
        pdf.cell(70, 6, txt='National Bank of Pakistan', align='C')
        pdf.cell(60, 6, txt='', align='C', ln=1)

        pdf.cell(60, 6, txt='Challan No.', align='L')
        pdf.cell(70, 6, txt='State Bank of Pakistan', align='C')
        pdf.cell(60, 6, txt='', align='C', ln=1)
        pdf.ln(4)
        pdf.cell(35, 6, txt='By whom Tender', align='L', border='LTR')
        pdf.cell(65, 6, txt='To be filled in the Remittes', align='C', border=1)
        pdf.cell(35, 6, txt='Amount', align='C', border='LTRB')
        pdf.cell(60, 6, txt='To be filled by the  the Treasury',
                 align='L', border='LTR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='Names or', align='C', border='LR')
        pdf.cell(35, 6, txt='Full Particulars ', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(60, 6, txt='Departmental Office of',
                 align='L',  border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='designation', align='C', border='LR')
        pdf.cell(35, 6, txt='of the ', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(60, 6, txt='the Treasury', align='L', border='LBR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='and address', align='C', border='LR')
        pdf.cell(35, 6, txt='remittance and', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='Head of', align='C', border='LR')
        pdf.cell(30, 6, txt='Order of', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='of the person', align='C', border='LR')
        pdf.cell(35, 6, txt='of authority (if', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='Account', align='C', border='LR')
        pdf.cell(30, 6, txt='the Bank', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='of whom behalf', align='C', border='LR')
        pdf.cell(35, 6, txt='any) ', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LBR')
        pdf.cell(30, 6, txt='money is paid', align='C', border='LBR')
        pdf.cell(35, 6, txt='', align='C', border='LBR')
        pdf.cell(35, 6, txt='', align='C', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR', ln=1)

        pdf.cell(35, 6, txt='Domicile Branch', align='L', border='LR')
        pdf.cell(30, 6, txt='District', align='C', border='LR')
        pdf.cell(35, 6, txt='Domicile Fee', align='C', border='LR')
        pdf.cell(35, 6, txt='{}'.format(
            self.total_fee), align='C', border='LR')
        pdf.cell(30, 6, txt='C-03806', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='DC Office, ICT', align='L', border='LR')
        pdf.cell(30, 6, txt='Magistrate', align='C', border='LR')
        pdf.cell(35, 6, txt='as per', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='Islamabad', align='L', border='LR')
        pdf.cell(30, 6, txt='Islamabad', align='C', border='LR')
        pdf.cell(35, 6, txt='attached list', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)

        pdf.cell(35, 6, txt='', align='L', border='LR')
        pdf.cell(30, 6, txt='{}'.format(challan_date), align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(35, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR')
        pdf.cell(30, 6, txt='', align='C', border='LR', ln=1)
        pdf.cell(35, 6, txt='', align='L', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR')
        pdf.cell(35, 6, txt='', align='C', border='LBR')
        pdf.cell(35, 6, txt='', align='C', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR')
        pdf.cell(30, 6, txt='', align='C', border='LBR', ln=1)
        pdf.set_font('arial', 'B', size=11)
        pdf.cell(35, 8, txt='', align='L', border='LBR')
        pdf.cell(30, 8, txt='', align='C', border='LBR')
        pdf.cell(35, 8, txt='Total', align='C', border='LBR')
        pdf.cell(35, 8, txt='{}'.format(
            self.total_fee), align='C', border='LBR')
        pdf.cell(30, 8, txt='', align='C', border='LBR')
        pdf.cell(30, 8, txt='', align='C', border='LBR', ln=1)

        pdf.set_font('arial', size=11)
        pdf.cell(65, 10, txt='In words Rupees ', align='R')
        pdf.set_font('arial', 'B', size=11)
        pdf.cell(0, 10, txt='{} Only'.format(
            num2words(self.total_fee)), align='L', ln=1)
        pdf.ln(5)
        pdf.set_font('arial', size=11)
        pdf.cell(50, 5, txt='Signature', align='LC')
        pdf.cell(50, 5, txt='Amount', align='C')
        pdf.cell(50, 5, txt='Received', align='C')
        pdf.cell(50, 5, txt='Treasury Officer', align='C', ln=1)

        pdf.cell(50, 5, txt='', align='LC')
        pdf.cell(50, 5, txt='', align='C')
        pdf.cell(50, 5, txt='Payment Treasury', align='C')
        pdf.cell(50, 5, txt='', align='C', ln=1)
        pdf.output('challan.pdf')
        path = 'challan.pdf'
        os.system(path)

    def edit_record(self, mod):
        if mod == 'edit' and self.rec_id == 0:
            return messagebox.showerror('Error', 'Nothing Selected')
        edit_window = Toplevel()
        edit_window.geometry('700x300')
        edit_window.title('Edit Record')
        dom_id = 0
        lbl_date = Label(edit_window, text='Domicile Date', width=15, font=self.btn_font).grid(
            row=1, column=0, padx=10, pady=10)
        inp_date = Entry(edit_window, width=15, font=self.btn_font)
        inp_date.grid(row=1, column=1, padx=10, pady=10)
        lbl_cnic = Label(edit_window, text='CNIC', width=15, font=self.btn_font).grid(
            row=1, column=2, padx=10, pady=10)
        inp_cnic = Entry(edit_window, width=15, font=self.btn_font)
        inp_cnic.grid(row=1, column=3, padx=10, pady=10)
        lbl_name = Label(edit_window, text='Applicant Name', width=15, font=self.btn_font).grid(
            row=2, column=0, padx=10, pady=10)
        inp_name = Entry(edit_window, width=15, font=self.btn_font)
        inp_name.grid(row=2, column=1, padx=10, pady=10)
        lbl_payment_type = Label(edit_window, text='Payment Type', width=15, font=self.btn_font).grid(
            row=2, column=2, padx=10, pady=10)
        inp_payment_type = Entry(edit_window, width=15, font=self.btn_font)
        inp_payment_type.grid(row=2, column=3, padx=10, pady=10)
        lbl_govt_fee = Label(edit_window, text='Govt Fee', width=15, font=self.btn_font).grid(
            row=3, column=0, padx=10, pady=10)
        inp_govt_fee = Entry(edit_window, width=15, font=self.btn_font)
        inp_govt_fee.grid(row=3, column=1, padx=10, pady=10)
        lbl_duplicate = Label(edit_window, text='Duplicate Entry', width=15, font=self.btn_font).grid(
            row=3, column=2, padx=10, pady=10)
        inp_duplicate = Entry(edit_window, width=15, font=self.btn_font)
        inp_duplicate.grid(row=3, column=3, padx=10, pady=10)
        lbl_app_type = Label(edit_window, text='Application Type', width=15, font=self.btn_font).grid(
            row=4, column=0, padx=10, pady=10)
        inp_app_type = Entry(edit_window, width=15, font=self.btn_font)
        inp_app_type.grid(
            row=4, column=1, padx=10, pady=10)
        lbl_Domicile_No = Label(edit_window, text='Domicile No', width=15, font=self.btn_font).grid(
            row=4, column=2, padx=10, pady=10)
        inp_Domicile_No = Entry(edit_window, width=15, font=self.btn_font)
        inp_Domicile_No.grid(row=4, column=3, padx=10, pady=10)
        lbl_req_type = Label(edit_window, text='Request Type', width=15, font=self.btn_font).grid(
            row=5, column=0, padx=10, pady=10)
        inp_req_type = Entry(edit_window, width=15, font=self.btn_font)
        inp_req_type.grid(
            row=5, column=1, padx=10, pady=10)
        if mod == 'edit':
            dom_id = 0
            Query = "Select * from cash_report Where Domicile_ID = %s;"
            con, cur = open_con(True)
            cur.execute(Query, [self.rec_id])
            self.agr_data = cur.fetchall()
            cur.close()
            con.close()
            dom_id = self.rec_id
            for row in self.agr_data:
                inp_date.insert(0, str(row['Domicile_Date']))
                inp_cnic.insert(0, str(row['cnic']))
                inp_name.insert(0, str(row['Applicant_Name']))
                inp_payment_type.insert(0, str(row['Payment_Type']))
                inp_govt_fee.insert(0, str(row['Govt_Fee']))
                inp_duplicate.insert(0, str(row['Duplicate_Entry']))
                inp_app_type.insert(0, str(row['Application_Type']))
                inp_Domicile_No.insert(0, str(row['Domicile_No']))
                inp_req_type.insert(0, str(row['Request_Type']))
        def update_record(mod):
            widget_list = [inp_date.get(), inp_name.get(), inp_cnic.get(), inp_payment_type.get(
            ), inp_govt_fee.get(), inp_app_type.get(), inp_duplicate.get(), inp_Domicile_No.get()]
            for item in widget_list:
                if item.lower().find('update') != -1 or item.lower().find('delete') != -1 or item.lower().find('char(') != -1:
                    return messagebox.showerror('Error', 'One of Input widgets contain miliciouse word e.g "delete", "update"')
            if mod == 'new':
                Query = "Insert into Cash_Report (Domicile_Date, Applicant_Name, cnic, Payment_Type, Govt_Fee, Application_Type, Duplicate_Entry, Domicile_No, Request_Type) values (%s, %s, %s, %s, %s, %s, %s, %s, %s);"
                parm_list = [inp_date.get(), inp_name.get(), inp_cnic.get(), inp_payment_type.get(), inp_govt_fee.get(), inp_app_type.get(), inp_duplicate.get(), inp_Domicile_No.get(), inp_req_type.get()]
            else:
                Query = "Update Cash_Report Set Domicile_Date=%s, Applicant_Name = %s, cnic=%s, Payment_Type = %s, Govt_Fee = %s, Application_Type=%s, Duplicate_Entry=%s, Domicile_No=%s, Request_Type=%s Where Domicile_ID = %s;"
                parm_list = [inp_date.get(), inp_name.get(), inp_cnic.get(), inp_payment_type.get(), inp_govt_fee.get(), inp_app_type.get(), inp_duplicate.get(), inp_Domicile_No.get(), inp_req_type.get(),dom_id]
            con, cur = open_con(False)
            cur.execute(Query, parm_list)
            con.commit()
            cur.close()
            con.close()
            self.status_label.delete(0, 'end')
            self.status_label.insert(0, 'Record Updated')
            messagebox.showinfo("Success", "Record Updated")
            edit_window.destroy()

        btn_update = Button(edit_window, width=10, text='Save', font=self.btn_font, command=lambda:update_record(mod)).grid(
            row=6, column=0, columnspan=2, padx=10, pady=10)
        btn_exit = Button(edit_window, width=10, text='Exit', font=self.btn_font, command=edit_window.destroy).grid(
            row=6, column=2, columnspan=2, padx=10, pady=10)

    def run(self):
        self.__root.mainloop()


if __name__ == '__main__':
    obj = Cash_Report()
    obj.run()
